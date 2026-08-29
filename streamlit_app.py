import io

import pandas as pd
import streamlit as st

from extractor import extract_text
from ai_extractor import extract_invoice_data


# Page configuration
st.set_page_config(
    page_title="Invoice Data Extractor",
    page_icon="📄",
    layout="wide"
)


# Custom styling
st.markdown(
    """
    <style>
    div.stButton > button {
        background-color: #2563eb;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 600;
    }

    div.stButton > button:hover {
        background-color: #1d4ed8;
        color: white;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# Header
st.title("Invoice Data Extractor")

st.write(
    "Upload one or multiple invoices and automatically "
    "extract their information."
)


# File upload
uploaded_files = st.file_uploader(
    "Upload invoices",
    type=["pdf", "png", "jpg", "jpeg"],
    accept_multiple_files=True,
    help="Supported formats: PDF, PNG, JPG and JPEG"
)


# Helper function
def display_value(value):
    if value is None or value == "":
        return "Not found"

    if isinstance(value, list):
        return "\n".join(str(item) for item in value)

    return value

def normalize_value(value):
    if value is None:
        return None

    if isinstance(value, list):
        return ", ".join(
            str(item) for item in value
        )

    return value

if uploaded_files:

    st.success(
        f"{len(uploaded_files)} invoice(s) uploaded"
    )

    if st.button(
        "Extract invoice data",
        type="primary"
    ):

        all_invoice_data = []

        progress_bar = st.progress(0)

        status_text = st.empty()


        # Process invoices
        for index, uploaded_file in enumerate(
            uploaded_files
        ):

            status_text.write(
                f"Processing **{uploaded_file.name}**..."
            )

            try:

                # OCR / text extraction
                text = extract_text(
                    uploaded_file
                )

                if not text.strip():

                    st.warning(
                        f"No text could be extracted from "
                        f"{uploaded_file.name}"
                    )

                    progress_bar.progress(
                        (index + 1) / len(uploaded_files)
                    )

                    continue


                # AI integration
                invoice_data = extract_invoice_data(
                    text
                )


                # Add original filename
                invoice_data["filename"] = (
                    uploaded_file.name
                )


                all_invoice_data.append(
                    invoice_data
                )


                progress_bar.progress(
                    (index + 1) / len(uploaded_files)
                )

            except Exception as e:

                st.error(
                    f"Error processing "
                    f"{uploaded_file.name}: {e}"
                )

                progress_bar.progress(
                    (index + 1) / len(uploaded_files)
                )


        status_text.empty()


        # Stop if nothing was processed
        if not all_invoice_data:

            st.error(
                "No invoices could be processed."
            )

            st.stop()


        # Processing summary
        processed_count = len(
            all_invoice_data
        )

        failed_count = (
            len(uploaded_files)
            - processed_count
        )


        st.divider()

        st.subheader("Processing Summary")


        col1, col2 = st.columns(2)

        with col1:

            st.metric(
                "Successfully processed",
                f"{processed_count}/{len(uploaded_files)}"
            )

        with col2:

            st.metric(
                "Failed",
                failed_count
            )


        # Display each invoice
        for invoice_data in all_invoice_data:

            filename = invoice_data.get(
                "filename",
                "Unknown file"
            )

            supplier = invoice_data.get(
                "supplier",
                {}
            )

            customer = invoice_data.get(
                "customer",
                {}
            )

            invoice = invoice_data.get(
                "invoice",
                {}
            )

            items = invoice_data.get(
                "items",
                []
            )

            totals = invoice_data.get(
                "totals",
                {}
            )


            # Find missing important fields
            missing_fields = []


            if not supplier.get("name"):
                missing_fields.append(
                    "Supplier name"
                )

            if not supplier.get("eik"):
                missing_fields.append(
                    "Supplier EIK / BULSTAT"
                )

            if not supplier.get("iban"):
                missing_fields.append(
                    "Supplier IBAN"
                )

            if not invoice.get("number"):
                missing_fields.append(
                    "Invoice number"
                )

            if not invoice.get("date"):
                missing_fields.append(
                    "Invoice date"
                )

            if totals.get("total") is None:
                missing_fields.append(
                    "Total"
                )


            # Invoice section
            st.divider()

            st.subheader(
                f"📄 {filename}"
            )


            # Extraction status
            if missing_fields:

                st.warning(
                    "Missing information: "
                    + ", ".join(missing_fields)
                )

            else:

                st.success(
                    "All important fields were extracted."
                )


            # Supplier / Customer
            col1, col2 = st.columns(2)


            with col1:

                st.markdown(
                    "### Supplier"
                )

                st.write(
                    "**Name:**",
                    display_value(
                        supplier.get("name")
                    )
                )

                st.write(
                    "**EIK / BULSTAT:**",
                    display_value(
                        supplier.get("eik")
                    )
                )

                st.write(
                    "**VAT Number:**",
                    display_value(
                        supplier.get("vat_number")
                    )
                )

                st.write(
                    "**Address:**",
                    display_value(
                        supplier.get("address")
                    )
                )

                st.write(
                    "**IBAN:**",
                    display_value(
                        supplier.get("iban")
                    )
                )

                st.write(
                    "**BIC:**",
                    display_value(
                        supplier.get("bic")
                    )
                )


            with col2:

                st.markdown(
                    "### Customer"
                )

                st.write(
                    "**Name:**",
                    display_value(
                        customer.get("name")
                    )
                )

                st.write(
                    "**EIK / BULSTAT:**",
                    display_value(
                        customer.get("eik")
                    )
                )

                st.write(
                    "**VAT Number:**",
                    display_value(
                        customer.get("vat_number")
                    )
                )

                st.write(
                    "**Address:**",
                    display_value(
                        customer.get("address")
                    )
                )


            # Invoice information
            st.markdown(
                "### Invoice"
            )


            col1, col2, col3, col4 = st.columns(4)


            with col1:

                st.metric(
                    "Invoice Number",
                    display_value(
                        invoice.get("number")
                    )
                )


            with col2:

                st.metric(
                    "Date",
                    display_value(
                        invoice.get("date")
                    )
                )


            with col3:

                st.metric(
                    "Due Date",
                    display_value(
                        invoice.get("due_date")
                    )
                )


            with col4:

                st.metric(
                    "Currency",
                    display_value(
                        invoice.get("currency")
                    )
                )


            # Invoice items
            st.markdown(
                "### Items"
            )


            if items:

                items_table = pd.DataFrame(
                    items
                )

                items_table = items_table.rename(
                    columns={
                        "description": "Product / Service",
                        "quantity": "Quantity",
                        "unit_price": "Unit Price",
                        "vat_rate": "VAT %",
                        "total_price": "Total"
                    }
                )

                st.dataframe(
                    items_table,
                    use_container_width=True,
                    hide_index=True
                )

            else:

                st.info(
                    "No invoice items found."
                )


            # Totals
            st.markdown(
                "### Totals"
            )


            col1, col2, col3 = st.columns(3)


            with col1:

                st.metric(
                    "Subtotal",
                    display_value(
                        totals.get("subtotal")
                    )
                )


            with col2:

                st.metric(
                    "VAT",
                    display_value(
                        totals.get("vat")
                    )
                )


            with col3:

                st.metric(
                    "Total",
                    display_value(
                        totals.get("total")
                    )
                )


        # ==========================================
        # EXCEL EXPORT
        # ==========================================

        st.divider()

        st.subheader("Excel Export")


        # Prepare rows
        invoice_rows = []
        item_rows = []
        supplier_rows = []
        customer_rows = []


        for invoice_data in all_invoice_data:

            filename = invoice_data.get(
                "filename"
            )

            supplier = invoice_data.get(
                "supplier",
                {}
            )

            customer = invoice_data.get(
                "customer",
                {}
            )

            invoice = invoice_data.get(
                "invoice",
                {}
            )

            totals = invoice_data.get(
                "totals",
                {}
            )


            # Invoice
            invoice_rows.append({
                "File": filename,
                "Invoice Number": invoice.get(
                    "number"
                ),
                "Date": invoice.get(
                    "date"
                ),
                "Due Date": invoice.get(
                    "due_date"
                ),
                "Currency": invoice.get(
                    "currency"
                ),
                "Subtotal": totals.get(
                    "subtotal"
                ),
                "VAT": totals.get(
                    "vat"
                ),
                "Total": totals.get(
                    "total"
                )
            })


            # Supplier
            supplier_rows.append({
                "File": filename,
                "Name": normalize_value(
                    supplier.get("name")
                ),
                "EIK / BULSTAT": normalize_value(
                    supplier.get("eik")
                ),
                "VAT Number": normalize_value(
                    supplier.get("vat_number")
                ),
                "Address": normalize_value(
                    supplier.get("address")
                ),
                "IBAN": normalize_value(
                    supplier.get("iban")
                ),
                "BIC": normalize_value(
                    supplier.get("bic")
                )
            })


            # Customer
            customer_rows.append({
                "File": filename,
                "Name": customer.get(
                    "name"
                ),
                "EIK / BULSTAT": customer.get(
                    "eik"
                ),
                "VAT Number": customer.get(
                    "vat_number"
                ),
                "Address": customer.get(
                    "address"
                )
            })


            # Items
            for item in invoice_data.get(
                "items",
                []
            ):

                item_rows.append({
                    "File": filename,
                    "Invoice Number": invoice.get(
                        "number"
                    ),
                    "Product / Service": item.get(
                        "description"
                    ),
                    "Quantity": item.get(
                        "quantity"
                    ),
                    "Unit Price": item.get(
                        "unit_price"
                    ),
                    "VAT %": item.get(
                        "vat_rate"
                    ),
                    "Total": item.get(
                        "total_price"
                    )
                })


        # DataFrames
        invoices_df = pd.DataFrame(
            invoice_rows
        )

        items_df = pd.DataFrame(
            item_rows
        )

        suppliers_df = pd.DataFrame(
            supplier_rows
        )

        customers_df = pd.DataFrame(
            customer_rows
        )


        # ==========================================
        # CREATE EXCEL
        # ==========================================

        excel_file = io.BytesIO()


        from openpyxl.styles import (
            Font,
            PatternFill,
            Alignment,
            Border,
            Side
        )

        from openpyxl.worksheet.table import (
            Table,
            TableStyleInfo
        )

        from openpyxl.utils import (
            get_column_letter
        )


        with pd.ExcelWriter(
            excel_file,
            engine="openpyxl"
        ) as writer:

            # Write sheets
            invoices_df.to_excel(
                writer,
                sheet_name="Invoices",
                index=False
            )

            items_df.to_excel(
                writer,
                sheet_name="Items",
                index=False
            )

            suppliers_df.to_excel(
                writer,
                sheet_name="Suppliers",
                index=False
            )

            customers_df.to_excel(
                writer,
                sheet_name="Customers",
                index=False
            )


            workbook = writer.book


            # Excel colors / styles
            header_fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=11
            )

            header_alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            body_alignment = Alignment(
                vertical="center"
            )

            border = Border(
                bottom=Side(
                    style="thin",
                    color="D9E1F2"
                )
            )


            # Format worksheets
            for worksheet in workbook.worksheets:

                # Freeze first row
                worksheet.freeze_panes = "A2"


                # Hide gridlines
                worksheet.sheet_view.showGridLines = False


                # Header
                for cell in worksheet[1]:

                    cell.fill = header_fill

                    cell.font = header_font

                    cell.alignment = (
                        header_alignment
                    )


                # Header height
                worksheet.row_dimensions[
                    1
                ].height = 26


                # Body
                for row in worksheet.iter_rows(
                    min_row=2
                ):

                    for cell in row:

                        cell.alignment = (
                            body_alignment
                        )

                        cell.border = border


                # Column widths
                for column_index in range(
                    1,
                    worksheet.max_column + 1
                ):

                    column_letter = (
                        get_column_letter(
                            column_index
                        )
                    )

                    header = str(
                        worksheet.cell(
                            row=1,
                            column=column_index
                        ).value
                    )


                    # Custom widths
                    if header == "File":

                        width = 30

                    elif header == "Invoice Number":

                        width = 24

                    elif header == "Product / Service":

                        width = 35

                    elif header == "Address":

                        width = 40

                    elif header == "IBAN":

                        width = 32

                    elif header == "Name":

                        width = 30

                    elif header in [
                        "EIK / BULSTAT",
                        "VAT Number",
                        "BIC"
                    ]:

                        width = 22

                    elif header in [
                        "Date",
                        "Due Date"
                    ]:

                        width = 16

                    elif header in [
                        "Currency",
                        "Quantity",
                        "VAT %"
                    ]:

                        width = 14

                    elif header in [
                        "Subtotal",
                        "VAT",
                        "Total",
                        "Unit Price"
                    ]:

                        width = 16

                    else:

                        max_length = 0

                        for cell in worksheet[
                            get_column_letter(
                                column_index
                            )
                        ]:

                            if cell.value is not None:

                                max_length = max(
                                    max_length,
                                    len(
                                        str(
                                            cell.value
                                        )
                                    )
                                )

                        width = max(
                            max_length + 3,
                            14
                        )


                    worksheet.column_dimensions[
                        column_letter
                    ].width = min(
                        width,
                        45
                    )


                # Number formatting
                for row in worksheet.iter_rows(
                    min_row=2
                ):

                    for cell in row:

                        header = worksheet.cell(
                            row=1,
                            column=cell.column
                        ).value


                        # Money
                        if header in [
                            "Subtotal",
                            "VAT",
                            "Total",
                            "Unit Price"
                        ]:

                            if isinstance(
                                cell.value,
                                (int, float)
                            ):

                                cell.number_format = (
                                    '#,##0.00'
                                )


                        # Quantity
                        elif header == "Quantity":

                            if isinstance(
                                cell.value,
                                (int, float)
                            ):

                                cell.number_format = (
                                    '#,##0.##'
                                )


                        # VAT
                        elif header == "VAT %":

                            if isinstance(
                                cell.value,
                                (int, float)
                            ):

                                cell.number_format = (
                                    '0.##'
                                )


                # Center important columns
                center_columns = [
                    "Date",
                    "Due Date",
                    "Currency",
                    "Quantity",
                    "VAT %",
                    "Invoice Number"
                ]


                for column_index in range(
                    1,
                    worksheet.max_column + 1
                ):

                    header = worksheet.cell(
                        row=1,
                        column=column_index
                    ).value


                    if header in center_columns:

                        for row_index in range(
                            2,
                            worksheet.max_row + 1
                        ):

                            worksheet.cell(
                                row=row_index,
                                column=column_index
                            ).alignment = Alignment(
                                horizontal="center",
                                vertical="center"
                            )


                # Excel table
                if worksheet.max_row > 1:

                    last_cell = worksheet.cell(
                        row=worksheet.max_row,
                        column=worksheet.max_column
                    ).coordinate


                    table = Table(
                        displayName=(
                            f"Table{worksheet.title}"
                        ),
                        ref=f"A1:{last_cell}"
                    )


                    table_style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )


                    table.tableStyleInfo = (
                        table_style
                    )

                    worksheet.add_table(
                        table
                    )


        # Prepare Excel file
        excel_file.seek(0)


        # Download button
        st.download_button(
            label="Download Excel",
            data=excel_file,
            file_name="invoice_data.xlsx",
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            type="primary"
        )